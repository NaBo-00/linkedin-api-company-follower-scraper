from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
from linkedin_api import Linkedin
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
import getpass

#Author and Copyright references
print("[Notice] - Follower List Scraper by NaBo-00\n[Notice] - Copyright NaBo-00 | All Rights Reserved")

#Other notice
print("[Notice] - Prerequisite: Admin access to the company's page")

print("[START] - linkedin_follower_scraper.py")

#Login to your LinkedIn Account -> get user credentials (profile needs access to a company's profile)
usr = input("[Info] - LinkedIn Credentials \n[Input] - LinkedIn Profile E-Mail: ")
pw = getpass.getpass("[Input] - Password: ")

#Insert company profile ID
comp_id = input("[Input] - Company ID: ")

#Specify search keywords
search_input = []
search_user = input("[Input] - Specify a search keyword: ")
search_input.append(search_user)

#Recursivly query for keywords
def add_search():
    add_keyword = input("[Input] - Do you want to add another search keyword? Type '0' for NO and '1' for YES: ")
    if add_keyword == "0":
        print("[Info] - Scraping process has started, please wait")
    elif add_keyword == "1":
        search_user = input("[Input] - Specify another search keyword: ")
        search_input.append(search_user)
        add_search()
    else:
        print("[ERROR] - Wrong input - RETRY!")
        add_search()

add_search()

#Return an Array with all Keywords
print("[Info] - Your specified search keywords: ", search_input)

#Set up Driver for Google Chrome Browser
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])

#Specify the path to chromedriver.exe
s = Service('chromedriver.exe')
myDriver = webdriver.Chrome(service=s, options= options)

#Specify the website you want to scrape data from
myDriver.get("https://www.linkedin.com/login")

#Set timeout
timeout=15
time.sleep(5)

#Handle cookie alert - deny
cookies = WebDriverWait(myDriver, timeout).until(EC.presence_of_element_located((By.XPATH, '//*[@id="artdeco-global-alert-container"]/div/section/div/div[2]/button[1]' ))).click()

#Identify login input fields
username = WebDriverWait(myDriver, timeout).until(EC.presence_of_element_located((By.XPATH, '//*[@id="username"]' )))
password = WebDriverWait(myDriver, timeout).until(EC.presence_of_element_located((By.XPATH, '//*[@id="password"]' )))

#Make sure that the input fields are empty
username.clear()
password.clear()

#Enter username and password
username.send_keys(usr)
password.send_keys(pw)

#Click Login Button
loginBtn = WebDriverWait(myDriver, timeout).until(EC.presence_of_element_located((By.XPATH, '//*[@id="organic-div"]/form/div[3]/button'))).click()

#Handle "Add phone number" by skipping the popp up window
if myDriver.find_elements(By.XPATH, '//*[@id="ember455"]/button') == 0:
    WebDriverWait(myDriver, timeout).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ember455"]/button'))).click()

time.sleep(3)

#Redirect to the companiy's profile page
myDriver.get("https://www.linkedin.com/company/" + comp_id + "/admin/")

time.sleep(3)

#Open follower list
follower_access = myDriver.find_elements(By.XPATH, '/html/body/div[6]/div[3]/div/div[3]/div[2]/div[1]/div[3]/div/div[1]/div[2]/div/section/div[3]/div[1]/div/div[2]/div/div/div[2]/div/a')

#Check if admin rights exist
if len(follower_access) == 0:
    print("[ERROR] - no admin access to company's profile")
    exit()
else:
    WebDriverWait(myDriver, timeout).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[6]/div[3]/div/div[3]/div[2]/div[1]/div[3]/div/div[1]/div[2]/div/section/div[3]/div[1]/div/div[2]/div/div/div[2]/div/a'))).click()

time.sleep(3)

#Redirect to follower list of the company
myDriver.get("https://www.linkedin.com/company/" + comp_id + "/admin/analytics/followers/")
time.sleep(2)

#Open and show follower list --> in case of issues: try to respecify the XPATH Selector for follower_list
follower_list = WebDriverWait(myDriver, timeout).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[6]/div[3]/div/div[3]/div[2]/div[3]/div/div/main/section/table/tbody/tr[5]/td/button'))).click()
time.sleep(3)

#Loop through the follower list and scrape all profile Ids
follower_ul = myDriver.find_element(By.XPATH, "/html/body/div[3]/div/div/div[2]/div[2]/div[1]")
number_followers = len(follower_ul.find_elements(By.CLASS_NAME, 'org-view-page-followers-modal__table-row'))

#Define Arrays to structure and store the general LinkedIn company follower's profile data
arrayProfileLink = []
arrayProfileName = []
arrayProfileId = []
arrayCounter = []
count = 1

#Append scraped data to the aforementioned Arrays
for i in range(1, number_followers + 1):
    #Retreive index number of follower
    arrayCounter.append(count)

    #Retreive Name
    optionsName = myDriver.find_elements(by=By.XPATH, value="/html/body/div[3]/div/div/div[2]/div[2]/div[1]/div[" + str(i) + "]/div/div[1]/a/div/div[2]/div[1]")
    for option in optionsName:
        arrayProfileName.append(option.text)

    #Retreive Link Information
    optionsLink = myDriver.find_elements(by=By.XPATH, value="/ html / body / div[3] / div / div / div[2] / div[2] / div[1] / div[" + str(i) + "] / div / div[1] / a")
    for option in optionsLink:
        arrayProfileLink.append(option.get_attribute('href'))  #Followers Profile Link
        arrayProfileId.append(option.get_attribute('href').split("/")[4])  #Followers Profile Id
        count += 1

#Define Arrays to structure and store the detailed LinkedIn company follower's profile data
arrayHeadline = []
arrayCurrentCompany = []
arrayCurrentJob = []
#Certificates are deprecated
arrayCertificates = []
arraySkills = []

#Authenticate to the unofficial linkedin-api using any Linkedin account credentials
api = Linkedin(usr, pw)

#GET information of all scraped follower profiles (Two API GET requests) and append retreived data to the aforementioned Arrays
for follower_id in arrayProfileId:
    #Certificate deprecated
    #arrayInsideCertificates = []
    arrayInsideSkills = []

    profile = api.get_profile(follower_id)
    profile1 = api.get_profile_skills(follower_id)

    arrayHeadline.append(profile["headline"])
    arrayCurrentCompany.append(profile['experience'][0]["companyName"])
    arrayCurrentJob.append(profile['experience'][0]["title"])

    #certifications are deprecated
    #number_certification = len(profile["certifications"])
    #for i in range(number_certification):
    #    arrayInsideCertificates.append(profile["certifications"][i]["authority"] + " | " + profile["certifications"][i]["name"])
    #arrayCertificates.append(arrayInsideCertificates)

    #add placeholder to certification
    arrayCertificates.append("certificate request deprecated")

    number_skills = len(profile1)
    for i in range(number_skills):
        arrayInsideSkills.append(profile1[i]["name"])
    arraySkills.append(arrayInsideSkills)

#Close Browser
myDriver.close()

#Store data in excel for keyword comparison
#save data in a DataFrame using Pandas
df_profile = pd.DataFrame({'ID': arrayCounter, 'Profile Name': arrayProfileName, 'Profile Headline': arrayHeadline, 'Current Job Title': arrayCurrentJob,  'Current Company': arrayCurrentCompany, 'Certificates': arrayCertificates, 'Skills': arraySkills, 'LinkedIn Profile ID': arrayProfileId, 'LinkedIn Profile Link': arrayProfileLink })
number_rows = len(df_profile.index)

#Save data in an excel file
sheet_name = "LinkedIn Followers"

#Configure excel file
dataframe = {sheet_name: df_profile}

#Create excel file
excelPath = 'LinkedIn_follower_' + comp_id + '.xlsx'

#Set Writer
writer = pd.ExcelWriter(excelPath, engine='xlsxwriter')

#Loop through the sheets, write the data and save the file
for sheet_name in dataframe.keys():
    dataframe[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
writer.close()

#Load excel workbook
wb = xl.load_workbook("LinkedIn_follower_" + comp_id + ".xlsx")
sheet = wb["LinkedIn Followers"]

#Highligh matching cells
for i in range(1, number_rows + 2):
    for search_keyword in search_input:
        value_C = sheet.cell(column=3, row=i).value
        value_D = sheet.cell(column=4, row=i).value
        value_F = sheet.cell(column=6, row=i).value
        value_G = sheet.cell(column=7, row=i).value

        if search_keyword.lower() in value_C.lower():
            sheet.cell(column=3, row=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
        if search_keyword.lower() in value_D.lower():
            sheet.cell(column=4, row=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        if search_keyword.lower() in value_F.lower():
            sheet.cell(column=6, row=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        if search_keyword.lower() in value_G.lower():
            sheet.cell(column=7, row=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

wb.save("LinkedIn_follower_" + comp_id + ".xlsx")

#Return info to the user
print("[Info] - Creating Excel File: LinkedIn_follower_" + comp_id + ".xlsx")
print("[Done] - Finished scraping and comparing")
print("[STOP] - linkedin_follower_scraper.py")